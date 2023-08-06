# -*- coding: utf-8 -*-
#!/usr/bin/env python
"""
=========================================================
Command line script (:mod:`pksci.scripts.txtlog2xlsx`)
=========================================================

.. currentmodule:: pksci.scripts.txtlog2xlsx

Convert text logs generated with
:py:class:`~pksci.tools.datautils.TextLogGenerator` class
to Excel ``xlsx`` files.

.. seealso::

   :py:mod:`~pksci.scripts.analyze_txtlog`
        command-line script for text log analysis and filtering

.. code-block:: python

   > txtlog2xlsx --help
   usage: txtlog2xlsx [-h] [--add-fnum] [--fnum FNUM] [--outpath OUTPATH]
                      [--overwrite] [--verbose]
                      txtlogs [txtlogs ...]

   positional arguments:
     txtlogs

   optional arguments:
     -h, --help         show this help message and exit
     --add-fnum
     --fnum FNUM
     --outpath OUTPATH
     --overwrite
     --verbose

.. autofunction:: txtlog2xlsx

Examples
========

To use this script, simply pass it one or more text log files generated
with the :py:class:`~pksci.tools.datautils.TextLogGenerator`
class::

    > txtlog2xlsx Tdlog-2013-*.txt

As always, view the script's help to see all available options::

    > txtlog2xlsx --help

"""
from __future__ import absolute_import, print_function, division

import argparse
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.cell import get_column_letter

from ..tools import get_fpath

__all__ = ['txtlog2xlsx']


def txtlog2xlsx(txtlogs=[], add_fnum=False, fnum=None, outpath=os.getcwd(),
                overwrite=False, verbose=False):
    """Convert text log files to Excel xlsx file.

    Parameters
    ----------
    txtlogs : sequence
        list of one or more text log files generated with
        :py:class:`~pksci.tools.datautils.TextLogGenerator` class
    add_fnum : bool, optional
        append integer number to output file.
    fnum : {None, int}, optional
        append ``fnum`` to output file
    outpath : str, optional
        set output path
    overwrite : bool, optional
        overwrite existing xlsx files of the same name
    verbose : bool, optional
        show verbose output

    """

    for findex, f in enumerate(txtlogs, start=1):
        if os.path.isfile(f):
            lines = None
            with open(f) as txtlog:
                lines = txtlog.readlines()

            fname = os.path.splitext(f)[0]
            xlsxfpath, xlsxfname = get_fpath(fname=fname, ext='xlsx',
                                             outpath=outpath,
                                             overwrite=overwrite,
                                             add_fnum=add_fnum,
                                             fnum=fnum,
                                             include_fname=True,
                                             verbose=verbose)

            if xlsxfpath is not None:
                fields = lines[0].strip()
                fields = re.split('\s{2,}', fields)

                wb = Workbook()
                ws = wb.worksheets[0]
                ws.title = fname[:31]

                # write field names in columns of first row
                for col_idx, field in enumerate(fields):
                    col = get_column_letter(col_idx + 1)
                    ws.cell('%s%s' % (col, 1)).value = field

                # write field values for each column, row
                for row, line in enumerate(lines[2:], start=2):
                    cells = re.split('\s{2,}', line.strip())
                    for col_idx, cell in enumerate(cells):
                        col = get_column_letter(col_idx + 1)
                        ws.cell('%s%s' % (col, row)).value = cell

                print('writing xlsx file: {}'.format(xlsxfname))
                wb.save(filename=xlsxfpath)
            else:
                if not verbose:
                    print('file exists: {}'.format(xlsxfname))
                    print('use `--overwrite` command-line arg to overwrite')
                    print('skipping file: {}\n'.format(f))


def _argparser():
    parser = argparse.ArgumentParser()
    parser.add_argument('--add-fnum', action='store_true',
                        help='append intger number to output file')
    parser.add_argument('--fnum', default=None,
                        help='append ``fnum`` to output file '
                        '(default: %(default)s)')
    parser.add_argument('--outpath', default=os.getcwd(),
                        help='output path (default: %(default)s)')
    parser.add_argument('--overwrite', action='store_true',
                        help='overwrite existing xlsx files')
    parser.add_argument('--verbose', action='store_true',
                        help='verbose output')
    parser.add_argument('txtlogs', nargs='+', help='text log')
    return parser


def main():

    args = _argparser().parse_args()
    txtlog2xlsx(**vars(args))

if __name__ == '__main__':
    sys.exit(main())
